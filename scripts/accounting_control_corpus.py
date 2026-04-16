#!/usr/bin/env python3
"""
accounting_control_corpus.py
============================

Real-world accounting control for Paper 2, T2 "duplicate sums".

On the khipu side, the bidirectional overlap script found that 35% of
validated khipus contain at least one pair of sums totalling **identical**
summand sets (Jaccard = 1.0), and 62% at Jaccard >= 0.5. A sceptical
reviewer can object: "maybe real bookkeeping duplicates totals too."

This script measures the same statistic on the Enron Spreadsheet Corpus
(Hermans 2015, ~15,000 real professional spreadsheets extracted from the
SEC's Enron investigation). If real accounting duplicates sums rarely, the
khipu excess is a genuine structural signature.

Method
------
For every .xlsx / .xlsm / .xls workbook in the corpus:

1. Open the file with formulas preserved (openpyxl data_only=False for
   xlsx/xlsm; pre-converted .xls handled via libreoffice).
2. Scan every sheet; for each cell whose value is a formula, extract:
   - single-range SUM:   =SUM(A1:A10)
   - multi-range SUM:    =SUM(A1:A5, C1:C5)
   - direct addition:    =A1+A2+A3   (>= 2 summands)
   The set of cell references is the "summand set", keyed by
   "<sheet>!<coord>" so cross-sheet references stay distinct.
3. Group sums by sheet (the natural section boundary in spreadsheets).
   Within each sheet, for every pair (sum_a, sum_b), compute
       Jaccard(S_a, S_b) = |S_a & S_b| / |S_a | S_b|
4. Per workbook (document), keep:
       n_sums
       max_jaccard, mean_jaccard
       pairs_ge_0.5, pairs_eq_1.0
5. Aggregate across the corpus and compare directly with the khipu
   baseline from ascher_bidirectional_overlap.py.

Outputs
-------
    output/accounting_control/comparison.csv
    output/accounting_control/comparison_report.md
    stdout : corpus stats + side-by-side vs khipu.

Usage
-----
    python scripts/accounting_control_corpus.py \
        --corpus-dir data/enron_corpus/spreadsheets \
        --sample N            # optional, random sample size
        --seed 20260416
"""

from __future__ import annotations

import argparse
import html as html_lib
import multiprocessing as mp
import random
import re
import sys
import warnings
import zipfile
from itertools import combinations
from pathlib import Path
from typing import List, Optional, Set, Tuple

import pandas as pd

# Silence the flood of openpyxl cosmetic warnings (print areas, headers,
# conditional formatting) that each load_workbook emits for Enron files.
warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CORPUS = REPO_ROOT / "data" / "enron_corpus" / "spreadsheets"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "accounting_control"


# ---------------------------------------------------------------------------
# Formula parsing
# ---------------------------------------------------------------------------

# Single cell reference, optionally sheet-qualified. The sheet name may be
# single-quoted. Absolute markers ($) are stripped.
CELL_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][\w\.]*)!)?"
    r"\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)",
    re.IGNORECASE,
)

# A range like A1:A10 (or sheet-qualified Sheet1!A1:A10, or $A$1:$A$10).
RANGE_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][\w\.]*)!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+):"
    r"\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+)",
    re.IGNORECASE,
)

SUM_RE = re.compile(r"SUM\s*\(", re.IGNORECASE)
# For =A1+A2+A3 style: the whole formula must be +-separated cell refs.
ADD_ONLY_RE = re.compile(
    r"^=\s*(?:[+-]?\s*[A-Za-z0-9_$']+![A-Z$]+\$?\d+|"
    r"[+-]?\s*[A-Z$]+\$?\d+)"
    r"(?:\s*[+]\s*[A-Za-z0-9_$']+![A-Z$]+\$?\d+|"
    r"\s*[+]\s*[A-Z$]+\$?\d+)+\s*$",
    re.IGNORECASE,
)


def expand_range(c1: str, r1: int, c2: str, r2: int, sheet: str) -> List[str]:
    """Enumerate all cells in a rectangular range."""
    col1 = column_index_from_string(c1.upper())
    col2 = column_index_from_string(c2.upper())
    if col2 < col1:
        col1, col2 = col2, col1
    if r2 < r1:
        r1, r2 = r2, r1
    # Guard against pathological whole-column ranges like A:A (r1==r2 at
    # MAX_ROW 1048576); cap at 100000 cells to avoid OOM.
    if (col2 - col1 + 1) * (r2 - r1 + 1) > 100_000:
        return []
    refs: List[str] = []
    for col in range(col1, col2 + 1):
        letter = get_column_letter(col)
        for row in range(r1, r2 + 1):
            refs.append(f"{sheet}!{letter}{row}")
    return refs


def _clean_sheet(raw: Optional[str], default: str) -> str:
    if not raw:
        return default
    return raw.strip("'")


def parse_sum_formula(formula: str, cur_sheet: str) -> Optional[List[str]]:
    """If ``formula`` is a SUM-equivalent aggregation, return its summand
    cell references (sheet-qualified). Otherwise None.

    Accepts:
        =SUM(A1:A10)
        =SUM(A1:A10, C1:C5)
        =SUM(A1, A3, A5)
        =A1+A2+A3  (>= 2 summands, all cell refs)
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return None

    body = formula[1:]                  # strip leading '='
    refs: List[str] = []

    if SUM_RE.match(body.lstrip()):
        # Pull the outer SUM(...) args
        # Handle nested parentheses very conservatively: take the content
        # between the first '(' after SUM and its matching ')'.
        start = body.find("(")
        if start < 0:
            return None
        depth = 0
        end = -1
        for i in range(start, len(body)):
            if body[i] == "(":
                depth += 1
            elif body[i] == ")":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        if end < 0:
            return None
        args = body[start + 1:end]
        # Reject SUM(SUM(...), ...) nested aggregations — they're legit
        # sums of sums; count them with the outer refs we can recover.
        for part in _split_top_level_commas(args):
            part = part.strip()
            if not part:
                continue
            m_r = RANGE_RE.fullmatch(part)
            if m_r:
                sh = _clean_sheet(m_r.group("sheet"), cur_sheet)
                refs.extend(expand_range(
                    m_r.group("c1"), int(m_r.group("r1")),
                    m_r.group("c2"), int(m_r.group("r2")), sh))
                continue
            m_c = CELL_RE.fullmatch(part)
            if m_c:
                sh = _clean_sheet(m_c.group("sheet"), cur_sheet)
                refs.append(f"{sh}!{m_c.group('col').upper()}{m_c.group('row')}")
                continue
            # Anything else (literals, other functions) -> skip silently
        if len(refs) >= 2:
            return refs
        return None

    if ADD_ONLY_RE.match(formula):
        for m in CELL_RE.finditer(body):
            sh = _clean_sheet(m.group("sheet"), cur_sheet)
            refs.append(f"{sh}!{m.group('col').upper()}{m.group('row')}")
        if len(refs) >= 2:
            return refs

    return None


def _split_top_level_commas(s: str) -> List[str]:
    """Split on commas not enclosed by parentheses."""
    parts: List[str] = []
    depth = 0
    cur: List[str] = []
    for ch in s:
        if ch == "(":
            depth += 1
            cur.append(ch)
        elif ch == ")":
            depth -= 1
            cur.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(cur))
            cur = []
        else:
            cur.append(ch)
    if cur:
        parts.append("".join(cur))
    return parts


# ---------------------------------------------------------------------------
# Workbook analysis
# ---------------------------------------------------------------------------

# Direct XML formula scan — an xlsx is a zip of xml; formulas live inside
# <c r="..."><f ...>FORMULA</f>...</c>. Regex-scanning the raw xml is
# 10-30x faster than instantiating openpyxl cell objects per file.
_CELL_FORMULA_RE = re.compile(
    r'<c\s+[^>]*?r="([A-Z]+\d+)"[^>]*>\s*<f(?:\s[^>]*)?>([^<]*)</f>',
    re.DOTALL,
)
_SHEET_XML_RE = re.compile(
    r'^xl/worksheets/(sheet\d+)\.xml$'
)


def _extract_sums_xml(path: Path) -> Optional[List[dict]]:
    """Fast path: open the xlsx as a zip and regex-scan each sheet xml for
    SUM-equivalent formulas. Returns None if the file is not a valid zip
    (caller falls back to openpyxl)."""
    sums: List[dict] = []
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            for name in names:
                m = _SHEET_XML_RE.match(name)
                if not m:
                    continue
                sh_key = m.group(1)        # synthetic per-sheet key
                try:
                    raw = z.read(name)
                except Exception:
                    continue
                try:
                    xml = raw.decode("utf-8")
                except UnicodeDecodeError:
                    xml = raw.decode("utf-8", errors="replace")
                for fm in _CELL_FORMULA_RE.finditer(xml):
                    coord = fm.group(1)
                    formula_body = html_lib.unescape(fm.group(2))
                    if not formula_body:
                        continue
                    refs = parse_sum_formula("=" + formula_body, sh_key)
                    if refs:
                        sums.append({
                            "sheet":    sh_key,
                            "coord":    f"{sh_key}!{coord}",
                            "summands": frozenset(refs),
                        })
    except (zipfile.BadZipFile, OSError, KeyError):
        return None
    return sums


def _extract_sums_openpyxl(path: Path) -> List[dict]:
    """Slow but tolerant fallback using openpyxl."""
    sums: List[dict] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return sums
    try:
        for sh_name in wb.sheetnames:
            try:
                ws = wb[sh_name]
            except Exception:
                continue
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str) or not val.startswith("="):
                        continue
                    refs = parse_sum_formula(val, sh_name)
                    if refs:
                        sums.append({
                            "sheet":    sh_name,
                            "coord":    f"{sh_name}!{cell.coordinate}",
                            "summands": frozenset(refs),
                        })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return sums


def extract_sums(path: Path) -> List[dict]:
    """Fast-path xml scan with openpyxl fallback."""
    fast = _extract_sums_xml(path)
    if fast is not None:
        return fast
    return _extract_sums_openpyxl(path)


def jaccard(a: Set[str], b: Set[str]) -> float:
    if not a and not b:
        return 0.0
    return len(a & b) / len(a | b)


def analyze_workbook(path: Path) -> Optional[dict]:
    """Return per-document Jaccard stats, or None if <2 sums in any sheet."""
    sums = extract_sums(path)
    if len(sums) < 2:
        return {"doc": str(path), "n_sums": len(sums),
                "n_pairs": 0, "max_jaccard": None, "mean_jaccard": None,
                "pairs_ge_0_5": 0, "pairs_eq_1_0": 0, "sheets_with_sums": 0}

    # Pair within each sheet only (natural section boundary).
    by_sheet: dict = {}
    for s in sums:
        by_sheet.setdefault(s["sheet"], []).append(s)

    jaccards: List[float] = []
    eq1 = ge5 = 0
    sheets_used = 0
    for sh_sums in by_sheet.values():
        if len(sh_sums) < 2:
            continue
        sheets_used += 1
        for a, b in combinations(sh_sums, 2):
            j = jaccard(a["summands"], b["summands"])
            jaccards.append(j)
            if j >= 0.99:
                eq1 += 1
            if j >= 0.5:
                ge5 += 1

    if not jaccards:
        return {"doc": str(path), "n_sums": len(sums),
                "n_pairs": 0, "max_jaccard": None, "mean_jaccard": None,
                "pairs_ge_0_5": 0, "pairs_eq_1_0": 0,
                "sheets_with_sums": sheets_used}

    return {
        "doc": str(path),
        "n_sums": len(sums),
        "n_pairs": len(jaccards),
        "max_jaccard": max(jaccards),
        "mean_jaccard": float(sum(jaccards) / len(jaccards)),
        "pairs_ge_0_5": ge5,
        "pairs_eq_1_0": eq1,
        "sheets_with_sums": sheets_used,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def gather_workbooks(corpus_dir: Path) -> List[Path]:
    exts = {".xlsx", ".xlsm"}
    return sorted(p for p in corpus_dir.rglob("*") if p.suffix.lower() in exts)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--corpus-dir", type=Path, default=DEFAULT_CORPUS)
    ap.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument("--sample", type=int, default=0,
                    help="Random sample size (0 = full corpus)")
    ap.add_argument("--seed", type=int, default=20260416)
    ap.add_argument("--workers", type=int, default=mp.cpu_count())
    ap.add_argument("--chunksize", type=int, default=32,
                    help="imap_unordered chunk size (higher = less IPC overhead)")
    ns = ap.parse_args()

    ns.output_dir.mkdir(parents=True, exist_ok=True)
    if not ns.corpus_dir.is_dir():
        sys.exit(f"ERROR: corpus directory not found: {ns.corpus_dir}")

    print(f"[*] Scanning {ns.corpus_dir} ...")
    paths = gather_workbooks(ns.corpus_dir)
    print(f"[*] Found {len(paths)} xlsx/xlsm workbooks")

    if ns.sample and ns.sample < len(paths):
        random.Random(ns.seed).shuffle(paths)
        paths = paths[:ns.sample]
        print(f"[*] Random sample: {len(paths)} workbooks")

    # Parallel processing with a pool of processes
    import time as _time
    t0 = _time.monotonic()
    print(f"[*] Analyzing with {ns.workers} workers (chunksize={ns.chunksize}) ...")
    with mp.Pool(ns.workers) as pool:
        raw_results = []
        for i, res in enumerate(pool.imap_unordered(
                analyze_workbook, paths, chunksize=ns.chunksize), 1):
            if res is not None:
                raw_results.append(res)
            if i % 1000 == 0:
                elapsed = _time.monotonic() - t0
                rate = i / elapsed if elapsed > 0 else 0
                eta = (len(paths) - i) / rate if rate > 0 else 0
                print(f"    {i:,} / {len(paths):,} processed "
                      f"({rate:.0f}/s, ETA {eta/60:.1f} min)")

    df = pd.DataFrame(raw_results)
    csv_path = ns.output_dir / "comparison.csv"
    df.to_csv(csv_path, index=False)
    print(f"[+] wrote {csv_path} ({len(df)} rows)")

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    has_pairs = df[df["n_pairs"] > 0]
    n_eligible = len(has_pairs)
    if n_eligible == 0:
        print("[!] No workbooks had >= 2 within-sheet SUM formulas.")
        return 1

    mean_max = has_pairs["max_jaccard"].mean()
    median_max = has_pairs["max_jaccard"].median()
    p_eq1 = 100.0 * (has_pairs["pairs_eq_1_0"] >= 1).sum() / n_eligible
    p_ge5 = 100.0 * (has_pairs["max_jaccard"] >= 0.5).sum() / n_eligible
    p_ge8 = 100.0 * (has_pairs["max_jaccard"] >= 0.8).sum() / n_eligible

    khipu_mean_max = 0.585
    khipu_p_eq1 = 35
    khipu_p_ge5 = 62
    khipu_p_ge8 = 45

    print()
    print("=" * 72)
    print("ACCOUNTING CONTROL CORPUS — Enron spreadsheets")
    print("=" * 72)
    print(f"Workbooks analysed            : {len(df)}")
    print(f"Workbooks with >=2 same-sheet SUMs (eligible): {n_eligible}")
    print(f"Mean max_jaccard (eligible)   : {mean_max:.3f}")
    print(f"Median max_jaccard (eligible) : {median_max:.3f}")
    print(f"Docs with >=1 pair Jaccard=1.0: {p_eq1:.1f}%")
    print(f"Docs with max Jaccard >= 0.5  : {p_ge5:.1f}%")
    print(f"Docs with max Jaccard >= 0.8  : {p_ge8:.1f}%")

    print()
    print("=" * 72)
    print("DIRECT COMPARISON WITH KHIPU CORPUS")
    print("=" * 72)
    print(f"{'Metric':<42} {'Khipu':>12} {'Enron':>12} {'Ratio':>8}")
    print(f"{'Mean max_jaccard':<42} {khipu_mean_max:>12.3f} "
          f"{mean_max:>12.3f} {khipu_mean_max / max(mean_max, 1e-9):>8.2f}x")
    print(f"{'% docs with Jaccard = 1.0':<42} {khipu_p_eq1:>11.1f}% "
          f"{p_eq1:>11.1f}% {khipu_p_eq1 / max(p_eq1, 1e-9):>8.2f}x")
    print(f"{'% docs with Jaccard >= 0.5':<42} {khipu_p_ge5:>11.1f}% "
          f"{p_ge5:>11.1f}% {khipu_p_ge5 / max(p_ge5, 1e-9):>8.2f}x")

    # ------------------------------------------------------------------
    # Markdown report
    # ------------------------------------------------------------------
    if mean_max < 0.2 and p_eq1 < 5:
        decision = ("Khipu duplication signal is **strong**. Real accounting "
                    "spreadsheets do not duplicate totals at comparable rates.")
    elif mean_max < 0.4:
        decision = ("Khipu duplication signal is **moderate**. Khipu rates "
                    "significantly exceed Enron but duplication is not "
                    "absent from real bookkeeping.")
    else:
        decision = ("Khipu duplication signal is **weak**. Enron spreadsheets "
                    "also duplicate totals at comparable rates; the khipu "
                    "argument needs to be reframed.")

    lines = []
    lines.append("# Accounting Control Corpus (Enron spreadsheets)")
    lines.append("")
    lines.append(
        "Test: does the 35 % of khipus with Jaccard = 1.0 sum pairs reflect "
        "a structural khipu signature, or do real accounting spreadsheets "
        "duplicate totals at similar rates?"
    )
    lines.append("")
    lines.append(f"## Corpus: {len(df)} workbooks "
                 f"(Hermans Enron Spreadsheet Corpus, figshare 1221767)")
    lines.append(f"Eligible (>=2 same-sheet SUMs): {n_eligible}")
    lines.append("")
    lines.append("| Metric | Khipu corpus | Enron corpus | Ratio |")
    lines.append("|---|---|---|---|")
    lines.append(
        f"| Mean max_jaccard (per doc) | {khipu_mean_max:.3f} | "
        f"{mean_max:.3f} | {khipu_mean_max / max(mean_max, 1e-9):.2f}x |"
    )
    lines.append(
        f"| Median max_jaccard | 0.690 | {median_max:.3f} | - |"
    )
    lines.append(
        f"| Docs with >=1 pair at Jaccard = 1.0 | {khipu_p_eq1:.0f}% | "
        f"{p_eq1:.1f}% | {khipu_p_eq1 / max(p_eq1, 1e-9):.1f}x |"
    )
    lines.append(
        f"| Docs with max Jaccard >= 0.5 | {khipu_p_ge5:.0f}% | "
        f"{p_ge5:.1f}% | {khipu_p_ge5 / max(p_ge5, 1e-9):.1f}x |"
    )
    lines.append(
        f"| Docs with max Jaccard >= 0.8 | {khipu_p_ge8:.0f}% | "
        f"{p_ge8:.1f}% | {khipu_p_ge8 / max(p_ge8, 1e-9):.1f}x |"
    )
    lines.append("")
    lines.append("## Decision")
    lines.append("")
    lines.append(decision)

    (ns.output_dir / "comparison_report.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8")
    print(f"[+] wrote {ns.output_dir / 'comparison_report.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
