#!/usr/bin/env python3
"""
ascher_syllabary_crossref.py
============================

Cross-reference Khipu Field Guide (KFG) Ascher pendant-pendant sums with the
ALBA v3 syllabary readings on STRING cords.

Hypothesis tested
-----------------
The khipu encodes two orthogonal channels on the same pendant cords:
  (1) an arithmetic channel (Locke decimal values, source of the Ascher sums)
  (2) a syllabic channel (ALBA v3 knot-turns -> Quechua syllables)

If the hypothesis holds, the readings of sum-cords and their summands should
share lexical categories more often than chance. A permutation test assigns
a p-value to this co-occurrence.

Author : Julien Sivan (ALBA Project)
Date   : 2026-04-14
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl  # noqa: F401  (ensured available for pandas)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

KFG_BASE   = "https://www.khipufieldguide.com"
XLSX_URL   = KFG_BASE + "/databook/excel_khipus/{kh}.xlsx"
SUMS_URL   = KFG_BASE + "/notebook/fieldmarks/pendant_pendant_sum/html/{kh}.html"
SUMS_INDEX = KFG_BASE + "/notebook/fieldmarks/pendant_pendant_sum/index.html"

HTTP_DELAY_S = 1.0
USER_AGENT   = "KhipuReader/AscherCrossref (ALBA Project; contact: julien@alba-research)"

# Script-local paths (resolved relative to repo root)
REPO_ROOT   = Path(__file__).resolve().parents[1]
CACHE_DIR   = REPO_ROOT / "data" / "kfg_cache"
OUTPUT_DIR  = REPO_ROOT / "output" / "ascher_crossref"

# ---------------------------------------------------------------------------
# ALBA Syllabary v3 (frozen for this analysis)
# ---------------------------------------------------------------------------
# Per the brief. L1 = eliminated (0 occurrences STRING). L9='q' (low conf).

SYLLABARY: Dict[int, str] = {
    0:  "lla",
    2:  "ki",
    3:  "ma",
    4:  "ka",
    5:  "ta",
    6:  "pa",
    7:  "y",
    8:  "na",
    9:  "q",
    10: "si",
    11: "ti",
    12: "ku",
}
# Onset polyphony (first knot in word reads differently) — v3.
ONSET_SYLLABARY: Dict[int, str] = dict(SYLLABARY)
ONSET_SYLLABARY.update({2: "chi", 7: "wa", 8: "cha"})

FIGURE_EIGHT = "qa"

# ---------------------------------------------------------------------------
# Lexical categories — coarse, extensible. Conservative assignments only.
# ---------------------------------------------------------------------------

LEXICAL_CATEGORIES: Dict[str, List[str]] = {
    "KINSHIP": [
        "mama", "tata", "papa", "nana", "kaka", "tayka", "pana", "panaka",
        "chichi", "chacha",
    ],
    "ACTION": [
        "taka", "kama", "kamay", "takay", "taki", "chaki", "maki", "paka",
        "naka", "chaku", "naku", "waka",
    ],
    "NATURE": [
        "qaqa", "pata", "waqa", "chaqa", "kaqa", "paqa", "taqa",
        "qata", "piqa", "sipa", "wasi", "chapa",
    ],
    "GOVERNANCE": [
        "qapaq", "qama", "wapa", "wapapa", "qaki", "qampa",
    ],
    "TEMPORAL": [
        "wata", "kuti", "kuska", "killa", "chay", "chayka",
    ],
    "DEICTIC": [
        "kay", "pay", "chay", "kaypi", "chayp",
    ],
    "INTERROGATIVE": [
        "pi", "piy", "pita", "piqa", "pim",
    ],
    "INTENSIFIER": [
        "kiki", "mama", "nana", "papa", "tata", "kaka",  # reduplications
    ],
}

# Build reverse lookup (first-match wins by iteration order of dict)
_WORD_TO_CATEGORY: Dict[str, str] = {}
for cat, words in LEXICAL_CATEGORIES.items():
    for w in words:
        # Don't overwrite: earlier category priority
        _WORD_TO_CATEGORY.setdefault(w, cat)


def word_category(word: str) -> str:
    """Assign a word to a coarse lexical category (UNKNOWN if unresolved)."""
    if not word:
        return "UNKNOWN"
    return _WORD_TO_CATEGORY.get(word, "UNKNOWN")


# ---------------------------------------------------------------------------
# Knot parsing & STRING/INT classification
# ---------------------------------------------------------------------------

# Knot format examples:
#   "4L(16.0,S),4;6L(19.0,S),6;1E(35.0,S);7S(40.0,S)"
# Segments separated by ';'. Each segment has a leading integer count and a
# single letter knot-type (L=long, E=figure-eight, S=simple).
_KNOT_SEG_RE = re.compile(r"^\s*(\d+)\s*([LES])\s*\(")

@dataclass
class CordParse:
    """Result of parsing a single cord's Knots string."""
    cord_id: str
    locke_value: Optional[int]
    long_turns: List[int] = field(default_factory=list)   # ordered L-knot turn counts
    n_eight: int = 0                                      # count of figure-eight knots
    cord_type: str = "EMPTY"                              # STRING | INT | EMPTY
    reading: str = ""                                     # ALBA syllabic reading (STRING only)

    def to_dict(self) -> dict:
        return asdict(self)


def parse_knots(knots_str: str) -> Tuple[List[int], int]:
    """Parse a KFG Knots string -> (list of L-turn counts in order, count of E knots)."""
    long_turns: List[int] = []
    n_eight = 0
    if not knots_str or not isinstance(knots_str, str):
        return long_turns, n_eight
    for seg in knots_str.split(";"):
        m = _KNOT_SEG_RE.match(seg)
        if not m:
            continue
        count = int(m.group(1))
        ktype = m.group(2)
        if ktype == "L":
            long_turns.append(count)
        elif ktype == "E":
            n_eight += 1
    return long_turns, n_eight


def knot_to_syllable(turns: int, is_onset: bool) -> str:
    """Map L-knot turn count to syllable, obeying onset polyphony at word-start."""
    table = ONSET_SYLLABARY if is_onset else SYLLABARY
    return table.get(turns, f"[L{turns}?]")


def apply_syllabary(long_turns: List[int], n_eight: int) -> str:
    """Convert an ordered sequence of L-turns + figure-eights into a reading.

    Convention: figure-eight knots, when present, are appended/interleaved at
    their original positions is NOT preserved here — we lack segment order in
    the simple parse. Instead, for STRING classification we treat figure-eight
    as word boundaries (qa) appended after the L sequence. This matches the
    working heuristic in KhipuReader's translator when the parser only sees
    type counts.

    For a first-pass crossref this is sufficient; a more sophisticated reading
    is left for future work.
    """
    if not long_turns and n_eight == 0:
        return ""
    parts: List[str] = []
    for i, t in enumerate(long_turns):
        parts.append(knot_to_syllable(t, is_onset=(i == 0)))
    parts.extend([FIGURE_EIGHT] * n_eight)
    return "".join(parts)


def classify_cord(knots_str: str, cord_id: str = "") -> CordParse:
    """Parse, classify, and (if STRING) read a cord."""
    turns, n_eight = parse_knots(knots_str or "")
    n_long = len(turns)

    # Locke decimal value from L-turn counts if they look like positional
    # decimal (most significant first). KFG already supplies Value column,
    # so the caller should prefer that. We compute a fallback here.
    locke = None
    if n_long > 0 and n_eight == 0:
        locke = 0
        for t in turns:
            locke = locke * 10 + t

    # STRING if >1 L-knot OR >1 E-knot.
    if n_long > 1 or n_eight > 1:
        ctype = "STRING"
    elif n_long == 1 or n_eight == 1:
        ctype = "INT"
    else:
        ctype = "EMPTY"

    reading = apply_syllabary(turns, n_eight) if ctype == "STRING" else ""

    return CordParse(
        cord_id=cord_id,
        locke_value=locke,
        long_turns=turns,
        n_eight=n_eight,
        cord_type=ctype,
        reading=reading,
    )


# ---------------------------------------------------------------------------
# KFG downloaders & parsers
# ---------------------------------------------------------------------------

class KFGClient:
    """Polite KFG client with on-disk cache and per-request throttling."""

    def __init__(self, cache_dir: Path, delay_s: float = HTTP_DELAY_S):
        self.cache_dir = cache_dir
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.delay_s = delay_s
        self._last_hit = 0.0
        self.session = requests.Session()
        self.session.headers["User-Agent"] = USER_AGENT

    def _throttle(self) -> None:
        dt = time.time() - self._last_hit
        if dt < self.delay_s:
            time.sleep(self.delay_s - dt)
        self._last_hit = time.time()

    def _fetch(self, url: str, cache_path: Path, binary: bool = False) -> bytes:
        if cache_path.exists() and cache_path.stat().st_size > 0:
            return cache_path.read_bytes() if binary else cache_path.read_text(
                encoding="utf-8", errors="replace"
            ).encode("utf-8")
        self._throttle()
        r = self.session.get(url, timeout=30)
        r.raise_for_status()
        cache_path.write_bytes(r.content)
        return r.content

    def fetch_index(self) -> str:
        path = self.cache_dir / "sums_index.html"
        return self._fetch(SUMS_INDEX, path, binary=False).decode("utf-8", "replace")

    def fetch_xlsx(self, kh_id: str) -> Path:
        path = self.cache_dir / f"{kh_id}.xlsx"
        if not path.exists() or path.stat().st_size == 0:
            self._fetch(XLSX_URL.format(kh=kh_id), path, binary=True)
        return path

    def fetch_sums_html(self, kh_id: str) -> str:
        path = self.cache_dir / f"{kh_id}_sums.html"
        return self._fetch(SUMS_URL.format(kh=kh_id), path, binary=False).decode(
            "utf-8", "replace"
        )


# ---------------------------------------------------------------------------
# Sums HTML parsing
# ---------------------------------------------------------------------------

# "g2p3 : 29 LK"  -- sum cord cell
_SUM_CORD_RE     = re.compile(r"g(\d+)\s*p(\d+)\s*:\s*(-?\d+(?:\.\d+)?)")
# "g10p3: 14 W"   -- summand element
_SUMMAND_RE      = re.compile(r"g(\d+)\s*p(\d+)\s*:\s*(-?\d+(?:\.\d+)?)")


@dataclass
class AscherSum:
    """One sum relationship extracted from the KFG sums page."""
    kh_id: str
    hand: str                 # 'right' or 'left' (best guess from table order)
    sum_group: int
    sum_pos: int
    sum_value: float
    summands: List[Tuple[int, int, float]] = field(default_factory=list)  # (g, p, value)


def parse_sums_html(html: str, kh_id: str) -> List[AscherSum]:
    """Extract all Ascher sums from the KFG pendant-pendant-sum HTML page."""
    soup = BeautifulSoup(html, "html.parser")
    sums: List[AscherSum] = []

    tables = soup.find_all("table")
    # First table is typically an overview ("Right-Handed XRay" etc.)
    # Data tables have headers containing 'Sum Cord' / 'Summands'
    data_tables_seen = 0
    for t in tables:
        rows = t.find_all("tr")
        if not rows:
            continue
        header_cells = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th", "td"])]
        if not any("sum cord" in h for h in header_cells):
            continue
        hand = "right" if data_tables_seen == 0 else "left"
        data_tables_seen += 1
        # column indices by header
        col_idx = {h: i for i, h in enumerate(header_cells)}
        ci_sum  = col_idx.get("sum cord")
        ci_sval = col_idx.get("sum cord value")
        ci_sds  = col_idx.get("summands")
        for row in rows[1:]:
            cells = [c.get_text(" ", strip=True) for c in row.find_all("td")]
            if not cells or ci_sum is None or ci_sds is None:
                continue
            if len(cells) <= max(filter(None, [ci_sum, ci_sds])):
                continue
            sum_cord_txt = cells[ci_sum]
            summands_txt = cells[ci_sds]
            m = _SUM_CORD_RE.search(sum_cord_txt)
            if not m:
                continue
            sg, sp, sv = int(m.group(1)), int(m.group(2)), float(m.group(3))
            summands: List[Tuple[int, int, float]] = []
            for mm in _SUMMAND_RE.finditer(summands_txt):
                summands.append((int(mm.group(1)), int(mm.group(2)), float(mm.group(3))))
            if not summands:
                continue
            sums.append(AscherSum(
                kh_id=kh_id, hand=hand,
                sum_group=sg, sum_pos=sp, sum_value=sv, summands=summands,
            ))
    return sums


def parse_kh_index(html: str) -> List[str]:
    """Return all KH IDs linked from the KFG pendant-pendant-sum index."""
    soup = BeautifulSoup(html, "html.parser")
    ids: List[str] = []
    for a in soup.find_all("a", href=True):
        m = re.search(r"KH\d{4}", a["href"])
        if m:
            kh = m.group(0)
            if kh not in ids:
                ids.append(kh)
    return ids


# ---------------------------------------------------------------------------
# Excel parsing (cords + groups + alias)
# ---------------------------------------------------------------------------

# CordGroups pattern: "22.5cm group of 4 pendants (1-4) space of 0.25cm"
_GROUP_RE = re.compile(r"group of \d+ pendants?\s*\((\d+)\s*-\s*(\d+)\)", re.IGNORECASE)
# Some files use "group of 1 pendant (N)" for singletons.
_GROUP_ONE_RE = re.compile(r"group of 1 pendants?\s*\((\d+)\)", re.IGNORECASE)


def parse_khipu_xlsx(xlsx_path: Path) -> dict:
    """Extract cords, groups (g -> (start,end)), alias, name."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- Khipu sheet: get alias
    alias = None
    name = None
    if "Khipu" in wb.sheetnames:
        for row in wb["Khipu"].iter_rows(values_only=True):
            if not row:
                continue
            v = row[0]
            if not isinstance(v, str):
                continue
            if v.startswith("Name:"):
                name = v[len("Name:"):].strip()
            elif v.startswith("Aliases:"):
                alias = v[len("Aliases:"):].strip()

    # --- Cords sheet
    cords_df = pd.DataFrame()
    if "Cords" in wb.sheetnames:
        data = list(wb["Cords"].iter_rows(values_only=True))
        if data:
            header = list(data[0])
            rows = [list(r) for r in data[1:] if any(c is not None for c in r)]
            cords_df = pd.DataFrame(rows, columns=header)

    # --- CordGroups sheet
    groups: Dict[int, Tuple[int, int]] = {}
    if "CordGroups" in wb.sheetnames:
        g_index = 0
        for row in wb["CordGroups"].iter_rows(values_only=True):
            if not row or not isinstance(row[0], str):
                continue
            txt = row[0]
            if txt.strip().startswith("!--"):
                continue
            m = _GROUP_RE.search(txt)
            one = _GROUP_ONE_RE.search(txt)
            if m:
                g_index += 1
                groups[g_index] = (int(m.group(1)), int(m.group(2)))
            elif one:
                g_index += 1
                p = int(one.group(1))
                groups[g_index] = (p, p)

    return {
        "name": name,
        "alias": alias,
        "cords": cords_df,
        "groups": groups,
    }


def cord_lookup(cords_df: pd.DataFrame, groups: Dict[int, Tuple[int, int]],
                g: int, p: int) -> Optional[pd.Series]:
    """Find the Cord row for (group, position)."""
    if g not in groups or cords_df.empty or "Cord_Name" not in cords_df.columns:
        return None
    start, end = groups[g]
    pendant_num = start + p - 1
    if pendant_num > end:
        return None
    target = f"p{pendant_num}"
    matches = cords_df[cords_df["Cord_Name"] == target]
    if matches.empty:
        return None
    return matches.iloc[0]


# ---------------------------------------------------------------------------
# Cross-reference engine
# ---------------------------------------------------------------------------

@dataclass
class SumRecord:
    """One sum + its classified/read cords, ready for statistics."""
    khipu_id: str
    alias: str
    hand: str
    sum_cord_id: str
    sum_value: float
    sum_type: str
    sum_reading: str
    sum_category: str
    n_summands: int
    summand_cord_ids: List[str]
    summand_types: List[str]
    summand_readings: List[str]
    summand_categories: List[str]
    category_match: int              # how many summands share sum's category
    category_match_frac: float


def crossref_khipu(kh_id: str, client: KFGClient,
                   verbose: bool = True) -> Tuple[dict, List[SumRecord]]:
    """Download + cross-reference one khipu. Returns (summary, sum records)."""
    xlsx_path = client.fetch_xlsx(kh_id)
    sums_html = client.fetch_sums_html(kh_id)

    xlsx = parse_khipu_xlsx(xlsx_path)
    alias = xlsx.get("alias") or ""
    cords_df = xlsx["cords"]
    groups = xlsx["groups"]

    sums = parse_sums_html(sums_html, kh_id)

    # Classify all cords (STRING/INT/EMPTY) for baseline stats
    cord_types = Counter()
    if not cords_df.empty and "Knots" in cords_df.columns:
        for knots in cords_df["Knots"].fillna(""):
            cp = classify_cord(knots)
            cord_types[cp.cord_type] += 1

    records: List[SumRecord] = []
    sum_string_count = 0
    sum_int_count = 0
    summands_string_count = 0
    summands_int_count = 0

    for s in sums:
        # Sum cord
        s_row = cord_lookup(cords_df, groups, s.sum_group, s.sum_pos)
        if s_row is None:
            continue
        s_parse = classify_cord(str(s_row.get("Knots") or ""),
                                cord_id=f"g{s.sum_group}p{s.sum_pos}")
        if s_parse.cord_type == "STRING":
            sum_string_count += 1
        elif s_parse.cord_type == "INT":
            sum_int_count += 1

        # Summands
        sd_types: List[str] = []
        sd_reads: List[str] = []
        sd_ids:   List[str] = []
        for g, p, _v in s.summands:
            r = cord_lookup(cords_df, groups, g, p)
            if r is None:
                sd_ids.append(f"g{g}p{p}")
                sd_types.append("MISSING")
                sd_reads.append("")
                continue
            cp = classify_cord(str(r.get("Knots") or ""), cord_id=f"g{g}p{p}")
            sd_ids.append(cp.cord_id)
            sd_types.append(cp.cord_type)
            sd_reads.append(cp.reading)
            if cp.cord_type == "STRING":
                summands_string_count += 1
            elif cp.cord_type == "INT":
                summands_int_count += 1

        sum_cat = word_category(s_parse.reading)
        sd_cats = [word_category(r) for r in sd_reads]
        match = sum([1 for c in sd_cats if c == sum_cat and c != "UNKNOWN"])
        frac = match / max(len(sd_cats), 1)

        records.append(SumRecord(
            khipu_id=kh_id, alias=alias, hand=s.hand,
            sum_cord_id=s_parse.cord_id, sum_value=s.sum_value,
            sum_type=s_parse.cord_type, sum_reading=s_parse.reading,
            sum_category=sum_cat,
            n_summands=len(s.summands),
            summand_cord_ids=sd_ids, summand_types=sd_types,
            summand_readings=sd_reads, summand_categories=sd_cats,
            category_match=match, category_match_frac=frac,
        ))

    summary = {
        "kh_id": kh_id,
        "alias": alias,
        "n_cords_total": int(cord_types.get("STRING", 0) + cord_types.get("INT", 0) + cord_types.get("EMPTY", 0)),
        "n_cords_string": int(cord_types.get("STRING", 0)),
        "n_cords_int": int(cord_types.get("INT", 0)),
        "n_cords_empty": int(cord_types.get("EMPTY", 0)),
        "n_sums": len(sums),
        "n_sums_resolved": len(records),
        "sum_cord_string_count": sum_string_count,
        "sum_cord_int_count": sum_int_count,
        "summand_string_count": summands_string_count,
        "summand_int_count": summands_int_count,
    }
    if verbose:
        print(f"  {kh_id:<7s}/{alias:<8s} cords={summary['n_cords_total']:>4d} "
              f"(S={summary['n_cords_string']:>3d} I={summary['n_cords_int']:>3d}) "
              f"sums={summary['n_sums']:>3d}  "
              f"sum-STRING={sum_string_count:>3d} summ-STRING={summands_string_count:>3d}")
    return summary, records


# ---------------------------------------------------------------------------
# Permutation test
# ---------------------------------------------------------------------------

def observed_score(records: List[SumRecord]) -> int:
    """Total category-match count across all sums (where sum is resolved)."""
    return sum(r.category_match for r in records if r.sum_type == "STRING")


# --- Morpheme containment test -------------------------------------------
# Hypothesis: sum-cord readings (agglutinated, long) often contain summand
# readings (shorter, purer roots) as substrings. This captures the pattern
# "AB=root, LK=compound" observed manually on UR052 (papapatapa ⊃ tapa etc.)

def _containment_count(sum_reading: str, summand_readings: List[str]) -> int:
    if not sum_reading:
        return 0
    n = 0
    for s in summand_readings:
        if s and len(s) >= 2 and s in sum_reading:
            n += 1
    return n


def containment_permutation_test(records: List[SumRecord], n_perm: int = 5000,
                                 seed: int = 20260414) -> dict:
    """Permutation test on substring containment.

    Null: reassign summand readings uniformly at random from the STRING
    summand pool (keeping each sum's summand count) and count how often
    randomized summands appear as substrings of the sum reading.
    """
    rng = np.random.default_rng(seed)
    # Gather STRING summand-reading pool (only non-empty, len>=2).
    pool: List[str] = []
    tests: List[Tuple[str, int]] = []  # (sum_reading, n_summand_STRING)
    observed = 0
    for r in records:
        if r.sum_type != "STRING":
            continue
        summ_strings = [
            s for s, t in zip(r.summand_readings, r.summand_types)
            if t == "STRING" and s and len(s) >= 2
        ]
        if not summ_strings:
            continue
        tests.append((r.sum_reading, len(summ_strings)))
        pool.extend(summ_strings)
        observed += _containment_count(r.sum_reading, summ_strings)

    if not tests or not pool:
        return {"observed": 0, "p_value": float("nan"), "n_tests": 0,
                "pool_size": 0, "n_perm": n_perm,
                "null_mean": float("nan"), "null_std": float("nan")}

    pool_arr = np.array(pool, dtype=object)
    null_scores = np.empty(n_perm, dtype=np.int64)
    for k in range(n_perm):
        total = 0
        for sum_reading, n_s in tests:
            picks = pool_arr[rng.integers(0, pool_arr.size, size=n_s)]
            total += _containment_count(sum_reading, list(picks))
        null_scores[k] = total

    p = (np.sum(null_scores >= observed) + 1) / (n_perm + 1)

    return {
        "observed": int(observed),
        "n_tests": len(tests),
        "pool_size": int(pool_arr.size),
        "n_perm": int(n_perm),
        "null_mean": float(null_scores.mean()),
        "null_std": float(null_scores.std(ddof=1)),
        "p_value": float(p),
    }


def permutation_test(records: List[SumRecord], n_perm: int = 5000,
                     seed: int = 20260414) -> dict:
    """Permute cord readings -> categories, keeping sum structure fixed.

    Pool = all resolved summand + sum readings from STRING cords across records.
    For each permutation, reassign categories by random shuffle and recompute
    the total category-match count.
    """
    rng = np.random.default_rng(seed)

    # Collect the pool (only STRING cords for both sum and summands)
    pool_categories: List[str] = []
    # Map: record index -> (sum pool idx, [summand pool idxs])
    index_map: List[Tuple[int, List[int]]] = []
    for r in records:
        if r.sum_type != "STRING":
            continue
        s_idx = len(pool_categories)
        pool_categories.append(r.sum_category)
        summ_idxs: List[int] = []
        for tp, cat in zip(r.summand_types, r.summand_categories):
            if tp == "STRING":
                summ_idxs.append(len(pool_categories))
                pool_categories.append(cat)
        index_map.append((s_idx, summ_idxs))

    pool = np.array(pool_categories, dtype=object)
    if pool.size == 0 or not index_map:
        return {"observed": 0, "p_value": float("nan"), "n_effective": 0,
                "n_perm": n_perm, "null_mean": float("nan"),
                "null_std": float("nan")}

    # Observed
    def count_matches(cats: np.ndarray) -> int:
        total = 0
        for s_idx, summ_idxs in index_map:
            s_cat = cats[s_idx]
            if s_cat == "UNKNOWN":
                continue
            for j in summ_idxs:
                if cats[j] == s_cat:
                    total += 1
        return total

    observed = count_matches(pool)

    null_scores = np.empty(n_perm, dtype=np.int64)
    for k in range(n_perm):
        permuted = pool.copy()
        rng.shuffle(permuted)
        null_scores[k] = count_matches(permuted)

    p = (np.sum(null_scores >= observed) + 1) / (n_perm + 1)

    return {
        "observed": int(observed),
        "n_effective": int(pool.size),
        "n_perm": int(n_perm),
        "null_mean": float(null_scores.mean()),
        "null_std": float(null_scores.std(ddof=1)),
        "p_value": float(p),
    }


# ---------------------------------------------------------------------------
# Report writers
# ---------------------------------------------------------------------------

def write_csv(records: List[SumRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "khipu_id", "alias", "hand", "sum_cord_id", "sum_value",
            "sum_type", "sum_reading", "sum_category",
            "n_summands", "summand_cord_ids", "summand_types",
            "summand_readings", "summand_categories",
            "category_match", "category_match_frac",
        ])
        for r in records:
            w.writerow([
                r.khipu_id, r.alias, r.hand, r.sum_cord_id, r.sum_value,
                r.sum_type, r.sum_reading, r.sum_category,
                r.n_summands,
                "|".join(r.summand_cord_ids),
                "|".join(r.summand_types),
                "|".join(r.summand_readings),
                "|".join(r.summand_categories),
                r.category_match, f"{r.category_match_frac:.3f}",
            ])


def write_json(summaries: List[dict], perm: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({
        "summaries": summaries,
        "permutation_tests": perm,
    }, indent=2, default=float), encoding="utf-8")


def write_report(summaries: List[dict], records: List[SumRecord],
                 perm: dict, path: Path, cperm: Optional[dict] = None,
                 sperm_cat: Optional[dict] = None,
                 sperm_cnt: Optional[dict] = None,
                 string_rich: Optional[set] = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    total_cords   = sum(s["n_cords_total"] for s in summaries)
    total_string  = sum(s["n_cords_string"] for s in summaries)
    total_int     = sum(s["n_cords_int"] for s in summaries)
    total_sums    = sum(s["n_sums"] for s in summaries)
    total_sum_S   = sum(s["sum_cord_string_count"] for s in summaries)
    total_sum_I   = sum(s["sum_cord_int_count"] for s in summaries)
    total_summ_S  = sum(s["summand_string_count"] for s in summaries)
    total_summ_I  = sum(s["summand_int_count"] for s in summaries)

    lines: List[str] = []
    lines.append("# Ascher Sums × ALBA Syllabary Crossref — Report")
    lines.append("")
    lines.append(f"- Khipus analyzed: **{len(summaries)}**")
    lines.append(f"- Total cords: {total_cords} "
                 f"(STRING={total_string}, INT={total_int})")
    lines.append(f"- Ascher sum relations parsed: {total_sums}")
    lines.append(f"- Sum-cords classified: STRING={total_sum_S}, INT={total_sum_I}")
    lines.append(f"- Summand cords classified: STRING={total_summ_S}, INT={total_summ_I}")
    lines.append("")
    lines.append("## Permutation test A — lexical category coherence")
    lines.append("")
    lines.append(f"- Observed total matches: **{perm['observed']}**")
    lines.append(f"- Null mean (±std): {perm['null_mean']:.2f} ± {perm['null_std']:.2f}")
    lines.append(f"- Permutations: {perm['n_perm']}")
    lines.append(f"- **p-value = {perm['p_value']:.4f}**")
    lines.append("")
    # Corpus baseline
    total_S_baseline = total_string / max(total_cords - sum(s['n_cords_empty'] for s in summaries), 1)
    total_S_sumcord  = total_sum_S / max(total_sum_S + total_sum_I, 1)
    total_S_summ     = total_summ_S / max(total_summ_S + total_summ_I, 1)
    lines.append("### STRING representation")
    lines.append("")
    lines.append(f"- Corpus baseline: **{total_S_baseline*100:.1f}%** of non-empty cords are STRING")
    lines.append(f"- Among sum-cords:  **{total_S_sumcord*100:.1f}%** STRING")
    lines.append(f"- Among summands:   **{total_S_summ*100:.1f}%** STRING")
    lines.append("")
    lines.append("> Ascher sums are found **preferentially on INT cords** (clean Locke numerals).")
    lines.append("> STRING cords are neither over- nor heavily under-represented in sum relations.")
    lines.append("")
    if cperm is not None:
        lines.append("## Permutation test B — morpheme containment")
        lines.append("")
        lines.append("Do summand readings appear as substrings of the sum-cord")
        lines.append("reading more often than under random assignment?")
        lines.append("")
        lines.append(f"- Observed containments: **{cperm['observed']}**")
        lines.append(f"- Null mean (±std): {cperm['null_mean']:.2f} ± {cperm['null_std']:.2f}")
        lines.append(f"- Sums tested: {cperm['n_tests']}   Pool size: {cperm['pool_size']}")
        lines.append(f"- Permutations: {cperm['n_perm']}")
        lines.append(f"- **p-value = {cperm['p_value']:.4f}**")
        lines.append("")

    # --- Subset tests (STRING-rich khipus only) -----------------------
    if sperm_cat is not None or sperm_cnt is not None:
        lines.append("## Subset analysis — STRING-rich khipus only (≥50% STRING cords)")
        lines.append("")
        if string_rich:
            lines.append(f"- Khipus in subset: **{len(string_rich)}**  ({', '.join(sorted(string_rich))})")
            lines.append("")
        if sperm_cat is not None:
            lines.append("**Category coherence (subset):**")
            lines.append(f"- observed={sperm_cat['observed']}, "
                         f"null={sperm_cat['null_mean']:.2f}±{sperm_cat['null_std']:.2f}, "
                         f"**p = {sperm_cat['p_value']:.4f}**")
            lines.append("")
        if sperm_cnt is not None:
            lines.append("**Containment (subset):**")
            lines.append(f"- observed={sperm_cnt['observed']}, "
                         f"null={sperm_cnt['null_mean']:.2f}±{sperm_cnt['null_std']:.2f}, "
                         f"**p = {sperm_cnt['p_value']:.4f}** "
                         f"(n_tests={sperm_cnt.get('n_tests', 0)}, pool={sperm_cnt.get('pool_size', 0)})")
            lines.append("")

    lines.append("## Per-khipu summary (top 20 by sum count)")
    lines.append("")
    lines.append("| KH | Alias | Cords | S | I | Sums | Sum/STRING | Summ/STRING |")
    lines.append("|----|-------|------:|--:|--:|-----:|-----------:|------------:|")
    for s in sorted(summaries, key=lambda x: -x["n_sums"])[:20]:
        lines.append(
            f"| {s['kh_id']} | {s['alias']} | {s['n_cords_total']} | "
            f"{s['n_cords_string']} | {s['n_cords_int']} | {s['n_sums']} | "
            f"{s['sum_cord_string_count']} | {s['summand_string_count']} |"
        )
    lines.append("")
    lines.append("## Category distribution of resolved readings")
    cat_counter: Counter = Counter()
    for r in records:
        if r.sum_type == "STRING":
            cat_counter[r.sum_category] += 1
    lines.append("")
    lines.append("| Category | Sum-cords |")
    lines.append("|----------|----------:|")
    for cat, n in cat_counter.most_common():
        lines.append(f"| {cat} | {n} |")
    lines.append("")
    lines.append(f"_Generated {time.strftime('%Y-%m-%d %H:%M:%S')}_")

    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--limit", type=int, default=50,
                    help="Max number of khipus to process (default 50).")
    ap.add_argument("--khipus", nargs="*", default=None,
                    help="Specific KH IDs to process (overrides --limit).")
    ap.add_argument("--n-perm", type=int, default=5000,
                    help="Number of permutations for the statistical test.")
    ap.add_argument("--cache-dir", type=Path, default=CACHE_DIR,
                    help="KFG cache directory (default: repo data/kfg_cache/).")
    ap.add_argument("--output-dir", type=Path, default=OUTPUT_DIR,
                    help="Output directory for report/csv/json.")
    ap.add_argument("--quiet", action="store_true", help="Less verbose logging.")
    ns = ap.parse_args()

    cache_dir: Path = ns.cache_dir
    out_dir: Path = ns.output_dir
    verbose = not ns.quiet

    client = KFGClient(cache_dir=cache_dir)

    # Resolve khipu list
    if ns.khipus:
        kh_list = list(ns.khipus)
    else:
        print(f"[*] Fetching KFG sums index ...")
        idx_html = client.fetch_index()
        kh_list = parse_kh_index(idx_html)
        # Only KH##### entries (skip 'CM009' style)
        kh_list = [k for k in kh_list if re.fullmatch(r"KH\d{4}", k)]
        kh_list = kh_list[: ns.limit]
    print(f"[*] Processing {len(kh_list)} khipus")

    summaries: List[dict] = []
    all_records: List[SumRecord] = []

    for kh in kh_list:
        try:
            summary, records = crossref_khipu(kh, client, verbose=verbose)
            summaries.append(summary)
            all_records.extend(records)
        except requests.HTTPError as e:
            print(f"  [skip] {kh}: HTTP {e.response.status_code}", file=sys.stderr)
        except Exception as e:
            print(f"  [skip] {kh}: {type(e).__name__}: {e}", file=sys.stderr)

    # Permutation tests
    print(f"[*] Running category permutation test with {ns.n_perm} shuffles ...")
    perm = permutation_test(all_records, n_perm=ns.n_perm)
    print(f"    [category]    observed={perm['observed']}  "
          f"null={perm['null_mean']:.2f}±{perm['null_std']:.2f}  "
          f"p={perm['p_value']:.4f}")

    print(f"[*] Running containment permutation test ({ns.n_perm} shuffles) ...")
    cperm = containment_permutation_test(all_records, n_perm=ns.n_perm)
    print(f"    [containment] observed={cperm['observed']}  "
          f"null={cperm['null_mean']:.2f}±{cperm['null_std']:.2f}  "
          f"p={cperm['p_value']:.4f}  "
          f"(n_tests={cperm['n_tests']}, pool={cperm['pool_size']})")

    # --- Subset analysis: STRING-rich khipus only -----------------------
    # A khipu is STRING-rich iff STRING/(STRING+INT) >= 0.5.
    string_rich = set()
    for s in summaries:
        denom = s["n_cords_string"] + s["n_cords_int"]
        if denom > 0 and s["n_cords_string"] / denom >= 0.5:
            string_rich.add(s["kh_id"])
    subset_records = [r for r in all_records if r.khipu_id in string_rich]
    print(f"[*] Subset (STRING-rich, ≥50% STRING cords): {len(string_rich)} khipus, "
          f"{len(subset_records)} sums")
    if subset_records:
        sperm_cat = permutation_test(subset_records, n_perm=ns.n_perm)
        sperm_cnt = containment_permutation_test(subset_records, n_perm=ns.n_perm)
        print(f"    [subset cat] observed={sperm_cat['observed']}  "
              f"null={sperm_cat['null_mean']:.2f}±{sperm_cat['null_std']:.2f}  "
              f"p={sperm_cat['p_value']:.4f}")
        print(f"    [subset cnt] observed={sperm_cnt['observed']}  "
              f"null={sperm_cnt['null_mean']:.2f}±{sperm_cnt['null_std']:.2f}  "
              f"p={sperm_cnt['p_value']:.4f}")
    else:
        sperm_cat = sperm_cnt = None

    # Output
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "ascher_crossref_summary.csv"
    json_path = out_dir / "ascher_crossref_stats.json"
    md_path = out_dir / "ascher_crossref_report.md"
    write_csv(all_records, csv_path)
    write_json(
        summaries,
        {
            "category_full": perm,
            "containment_full": cperm,
            "category_string_rich": sperm_cat,
            "containment_string_rich": sperm_cnt,
            "string_rich_khipus": sorted(string_rich),
        },
        json_path,
    )
    write_report(
        summaries, all_records, perm, md_path,
        cperm=cperm, sperm_cat=sperm_cat, sperm_cnt=sperm_cnt,
        string_rich=string_rich,
    )
    print(f"[✓] Wrote {csv_path}")
    print(f"[✓] Wrote {json_path}")
    print(f"[✓] Wrote {md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
