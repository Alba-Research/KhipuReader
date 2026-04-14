"""
ascher
======

Structural layer for Khipu Field Guide (KFG) Ascher pendant-pendant sums.

This module is intentionally minimal and structural. It does NOT apply any
syllabary (that is `syllabary.py`'s job) and does NOT perform semantic
analysis. It exposes:

  - :class:`CordParse`       — knot-string decomposition (L / S / E counts
                                and decimal contributions)
  - :func:`parse_knots`      — parse a KFG knot string into a CordParse
  - :class:`KFGClient`       — polite HTTP client with on-disk cache
  - :class:`AscherSum`       — one pendant-pendant sum relation
  - :func:`parse_sums_html`  — extract AscherSum objects from a sum page
  - :func:`parse_kh_index`   — list all KH IDs from the index page
  - :class:`KhipuData`       — KFG xlsx contents (cords DataFrame + groups)
  - :func:`parse_khipu_xlsx` — load a KH{NNNN}.xlsx file
  - :func:`resolve_cord`     — map (group, position) to a Cords row

The higher-level :class:`AscherGraph` (Stage 1) will be added on top in a
subsequent commit.

Cached files live under ``KhipuReader/data/kfg_cache/`` by default, shared
with any other scripts using this module.
"""

from __future__ import annotations

import re
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl  # noqa: F401  (required by pandas.read_excel / openpyxl.load_workbook)

try:
    import networkx as nx
    HAS_NX = True
except ImportError:
    HAS_NX = False

# ---------------------------------------------------------------------------
# HTTP / cache configuration
# ---------------------------------------------------------------------------

KFG_BASE   = "https://www.khipufieldguide.com"
XLSX_URL   = KFG_BASE + "/databook/excel_khipus/{kh}.xlsx"
SUMS_URL   = KFG_BASE + "/notebook/fieldmarks/pendant_pendant_sum/html/{kh}.html"
SUMS_INDEX = KFG_BASE + "/notebook/fieldmarks/pendant_pendant_sum/index.html"

HTTP_DELAY_S = 1.0
USER_AGENT   = "KhipuReader/ascher (ALBA Project)"

# Default cache lives under the package root: <repo>/data/kfg_cache/
_PKG_ROOT   = Path(__file__).resolve().parents[2]
DEFAULT_CACHE_DIR = _PKG_ROOT / "data" / "kfg_cache"

# ---------------------------------------------------------------------------
# Knot parser
# ---------------------------------------------------------------------------
#
# KFG knot-string format
# ----------------------
# Segments are separated by ``;``. Each segment has the shape
#
#     {count}{type}({pos_cm},{ply}),{contrib}
#
# where
#   - ``count`` is an integer: turn count for an L-knot, cluster size for an
#     S-knot, always 1 for an E (figure-eight) knot.
#   - ``type`` is one of ``L``, ``S``, ``E``.
#   - ``pos_cm`` is the cm position on the cord.
#   - ``ply`` is an orientation letter (``S``/``Z``/``U``) — unused here.
#   - ``contrib`` is the integer decimal contribution of that segment to the
#     cord's Locke value. This field is optional in some E entries; when
#     absent we fall back to ``count`` (always 1 for E).

_SEG_RE = re.compile(
    r"(\d+)\s*([LES])\s*\(\s*([-\d.]+)\s*,\s*[^\)]*\)\s*(?:,\s*(\d+))?"
)


@dataclass
class CordParse:
    """Structural parse of a single cord's Knots string.

    No syllabary is applied here — see the caller's own syllabary module
    to turn ``long_turns`` into a reading.
    """

    knots_raw: str = ""

    # Per-type counts (number of segments, not sum of turns)
    n_long: int = 0
    n_simple: int = 0
    n_eight: int = 0

    # Ordered segment data
    long_turns: List[int] = field(default_factory=list)   # L-knot turn counts
    s_values: List[int] = field(default_factory=list)     # S-knot decimal contribs
    l_values: List[int] = field(default_factory=list)     # L-knot decimal contribs
    e_values: List[int] = field(default_factory=list)     # E-knot decimal contribs

    # Aggregated decimal totals (all drawn from the trailing ",N" field)
    s_value_total: int = 0       # the scribe's "padding" contribution
    l_value_total: int = 0       # the "word" contribution
    e_value_total: int = 0
    locke_value: int = 0         # S + L + E

    # Classification: STRING iff >1 L-knot or >1 E-knot; else INT; else EMPTY.
    cord_type: str = "EMPTY"


def parse_knots(knots_str: Optional[str]) -> CordParse:
    """Parse a KFG ``Knots`` string into a :class:`CordParse`.

    The returned object contains only structural information (counts,
    per-type values, classification). No syllabary is applied — the
    caller owns that decision.
    """
    cp = CordParse(knots_raw=knots_str or "")
    if not knots_str or not isinstance(knots_str, str):
        return cp

    for m in _SEG_RE.finditer(knots_str):
        count = int(m.group(1))
        typ = m.group(2)
        contrib = int(m.group(4)) if m.group(4) is not None else count
        if typ == "L":
            cp.n_long += 1
            cp.long_turns.append(count)
            cp.l_values.append(contrib)
            cp.l_value_total += contrib
        elif typ == "S":
            cp.n_simple += 1
            cp.s_values.append(contrib)
            cp.s_value_total += contrib
        elif typ == "E":
            cp.n_eight += 1
            cp.e_values.append(contrib)
            cp.e_value_total += contrib

    cp.locke_value = cp.s_value_total + cp.l_value_total + cp.e_value_total

    if cp.n_long > 1 or cp.n_eight > 1:
        cp.cord_type = "STRING"
    elif cp.n_long == 1 or cp.n_eight == 1 or cp.n_simple >= 1:
        cp.cord_type = "INT"
    else:
        cp.cord_type = "EMPTY"
    return cp


# ---------------------------------------------------------------------------
# KFG HTTP client (polite, on-disk cache)
# ---------------------------------------------------------------------------

class KFGClient:
    """Small HTTP client for khipufieldguide.com with file cache + throttling.

    Parameters
    ----------
    cache_dir : Path, optional
        Directory to store downloaded artefacts. Defaults to the shared
        ``<repo>/data/kfg_cache/`` directory.
    delay_s : float
        Minimum seconds between two successive HTTP requests (rate limiting).
    """

    def __init__(self, cache_dir: Optional[Path] = None,
                 delay_s: float = HTTP_DELAY_S):
        self.cache_dir = Path(cache_dir) if cache_dir else DEFAULT_CACHE_DIR
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.delay_s = delay_s
        self._last = 0.0
        self.session = requests.Session()
        self.session.headers["User-Agent"] = USER_AGENT

    def _throttle(self) -> None:
        dt = time.time() - self._last
        if dt < self.delay_s:
            time.sleep(self.delay_s - dt)
        self._last = time.time()

    def _fetch(self, url: str, cache_path: Path, binary: bool) -> bytes:
        if cache_path.exists() and cache_path.stat().st_size > 0:
            if binary:
                return cache_path.read_bytes()
            return cache_path.read_text(encoding="utf-8", errors="replace").encode("utf-8")
        self._throttle()
        r = self.session.get(url, timeout=30)
        r.raise_for_status()
        cache_path.write_bytes(r.content)
        return r.content

    def fetch_index(self) -> str:
        """Return the HTML of the pendant-pendant-sum index page."""
        return self._fetch(SUMS_INDEX, self.cache_dir / "sums_index.html",
                           False).decode("utf-8", "replace")

    def fetch_xlsx(self, kh: str) -> Path:
        """Download (or use cached) the KFG xlsx for ``kh`` and return its path."""
        path = self.cache_dir / f"{kh}.xlsx"
        if not path.exists() or path.stat().st_size == 0:
            self._fetch(XLSX_URL.format(kh=kh), path, True)
        return path

    def fetch_sums_html(self, kh: str) -> str:
        """Return the HTML of the pendant-pendant-sum page for ``kh``."""
        return self._fetch(SUMS_URL.format(kh=kh),
                           self.cache_dir / f"{kh}_sums.html",
                           False).decode("utf-8", "replace")


# ---------------------------------------------------------------------------
# Pendant-pendant sum parsing (HTML)
# ---------------------------------------------------------------------------

# "g2p3 : 29 LK"
_SUM_CORD_RE = re.compile(r"g(\d+)\s*p(\d+)\s*:\s*(-?\d+(?:\.\d+)?)")
# "g10p3: 14 W"
_SUMMAND_RE  = re.compile(r"g(\d+)\s*p(\d+)\s*:\s*(-?\d+(?:\.\d+)?)")


@dataclass
class AscherSum:
    """One pendant-pendant sum relation extracted from the KFG sums HTML."""

    kh_id: str
    hand: str                        # 'right' or 'left' (by table order)
    sum_group: int
    sum_pos: int
    sum_value: float
    summands: List[Tuple[int, int, float]] = field(default_factory=list)  # (g, p, value)


def parse_kh_index(html: str) -> List[str]:
    """Return the unique list of KH IDs referenced by the sums index page."""
    soup = BeautifulSoup(html, "html.parser")
    ids: List[str] = []
    for a in soup.find_all("a", href=True):
        m = re.search(r"KH\d{4}", a["href"])
        if m and m.group(0) not in ids:
            ids.append(m.group(0))
    return ids


def parse_sums_html(html: str, kh_id: str) -> List[AscherSum]:
    """Extract all pendant-pendant sums from the KFG HTML page for ``kh_id``.

    The page typically contains two data tables (right-handed then
    left-handed). Table order determines the ``hand`` attribute.
    """
    soup = BeautifulSoup(html, "html.parser")
    sums: List[AscherSum] = []
    data_seen = 0
    for t in soup.find_all("table"):
        rows = t.find_all("tr")
        if not rows:
            continue
        headers = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th", "td"])]
        if not any("sum cord" in h for h in headers):
            continue
        hand = "right" if data_seen == 0 else "left"
        data_seen += 1
        ci = {h: i for i, h in enumerate(headers)}
        ci_sum = ci.get("sum cord")
        ci_sds = ci.get("summands")
        if ci_sum is None or ci_sds is None:
            continue
        for row in rows[1:]:
            cells = [c.get_text(" ", strip=True) for c in row.find_all("td")]
            if len(cells) <= max(ci_sum, ci_sds):
                continue
            m = _SUM_CORD_RE.search(cells[ci_sum])
            if not m:
                continue
            sg, sp, sv = int(m.group(1)), int(m.group(2)), float(m.group(3))
            summands = [
                (int(mm.group(1)), int(mm.group(2)), float(mm.group(3)))
                for mm in _SUMMAND_RE.finditer(cells[ci_sds])
            ]
            if not summands:
                continue
            sums.append(AscherSum(kh_id, hand, sg, sp, sv, summands))
    return sums


# ---------------------------------------------------------------------------
# Khipu xlsx parsing
# ---------------------------------------------------------------------------

# CordGroups entries such as:
#   "22.5cm group of 4 pendants (1-4) space of 0.25cm"
_GROUP_RE  = re.compile(r"group of \d+ pendants?\s*\((\d+)\s*-\s*(\d+)\)", re.IGNORECASE)
_GROUP_ONE = re.compile(r"group of 1 pendants?\s*\((\d+)\)", re.IGNORECASE)


@dataclass
class KhipuData:
    """In-memory representation of a single KFG xlsx file."""

    kh_id: str
    alias: str
    cords: pd.DataFrame
    groups: Dict[int, Tuple[int, int]]         # group_number -> (p_start, p_end)
    cord_to_group: Dict[str, int]              # "p23" -> group_number


def parse_khipu_xlsx(path: Path, kh_id: str) -> KhipuData:
    """Load a KFG ``KH{NNNN}.xlsx`` file.

    Returns a :class:`KhipuData` holding the Cords DataFrame, the parsed
    CordGroups layout, and a convenience ``cord_to_group`` mapping.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    alias = ""
    if "Khipu" in wb.sheetnames:
        for row in wb["Khipu"].iter_rows(values_only=True):
            if row and isinstance(row[0], str) and row[0].startswith("Aliases:"):
                alias = row[0][len("Aliases:"):].strip()
                break

    cords_df = pd.DataFrame()
    if "Cords" in wb.sheetnames:
        data = list(wb["Cords"].iter_rows(values_only=True))
        if data:
            header = list(data[0])
            rows = [list(r) for r in data[1:] if any(c is not None for c in r)]
            cords_df = pd.DataFrame(rows, columns=header)

    groups: Dict[int, Tuple[int, int]] = {}
    if "CordGroups" in wb.sheetnames:
        gi = 0
        for row in wb["CordGroups"].iter_rows(values_only=True):
            if not row or not isinstance(row[0], str):
                continue
            if row[0].strip().startswith("!--"):
                continue
            m = _GROUP_RE.search(row[0])
            mo = _GROUP_ONE.search(row[0])
            if m:
                gi += 1
                groups[gi] = (int(m.group(1)), int(m.group(2)))
            elif mo:
                gi += 1
                p = int(mo.group(1))
                groups[gi] = (p, p)

    cord_to_group: Dict[str, int] = {}
    for g, (start, end) in groups.items():
        for p in range(start, end + 1):
            cord_to_group[f"p{p}"] = g

    return KhipuData(
        kh_id=kh_id, alias=alias, cords=cords_df, groups=groups,
        cord_to_group=cord_to_group,
    )


def resolve_cord(kd: KhipuData, g: int, p: int) -> Optional[pd.Series]:
    """Return the Cords-sheet row for group ``g`` / within-group position ``p``.

    Position semantics
    ------------------
    We use the pendant number (``pn``) itself — extracted from the ``p{N}``
    name — as the physical-position proxy on the primary cord. This avoids
    any dependence on xlsx row ordering.

    Returns ``None`` if the group is unknown, the position exceeds the
    group's range, or the resulting cord name cannot be found in the
    Cords sheet.
    """
    if g not in kd.groups:
        return None
    start, end = kd.groups[g]
    pn = start + p - 1
    if pn > end:
        return None
    name = f"p{pn}"
    if kd.cords.empty or "Cord_Name" not in kd.cords.columns:
        return None
    hit = kd.cords[kd.cords["Cord_Name"] == name]
    if hit.empty:
        return None
    row = hit.iloc[0].copy()
    row["_cord_name"] = name
    row["_group"] = g
    row["_index"] = pn
    return row


# ---------------------------------------------------------------------------
# AscherGraph — structural summary of a khipu's pendant-pendant sums
# ---------------------------------------------------------------------------

# Role constants exposed via AscherGraph.get_role()
ROLE_HEADER = "HEADER"   # sum-cord (carries the section total; often a word)
ROLE_DATA   = "DATA"     # summand only
ROLE_BOTH   = "BOTH"     # cascade intermediate: sum-cord of one sum, summand of another
ROLE_FREE   = "FREE"     # not involved in any sum


@dataclass
class AscherCordAnnotation:
    """Per-cord Ascher annotation attached to a :class:`CordTranslation`.

    All fields default to ``None`` / empty so a cord that does not
    participate in any sum is an identifiable no-op. Consumers (exporters,
    report writers) must treat a ``None`` ``CordTranslation.ascher`` as
    "no Ascher information" — v1 behaviour.
    """

    role: str                                            # HEADER / DATA / BOTH / FREE
    cascade_depth: int = 0
    s_value_total: int = 0                               # scribe padding contribution
    # HEADER / BOTH -------------------------------------------------------
    expected_locke: Optional[float] = None               # KFG-declared sum total
    sum_hand: Optional[str] = None                       # 'right' | 'left'
    summand_pendant_names: List[str] = field(default_factory=list)
    # DATA / BOTH ---------------------------------------------------------
    referenced_by_pendant_names: List[str] = field(default_factory=list)
    # Pass 3B -------------------------------------------------------------
    verified: Optional[bool] = None                      # Locke parse == expected
    decomposition: Optional[str] = None                  # morpho decomposition (3B)
    # Pass 3A -------------------------------------------------------------
    reclassified_cord_type: Optional[str] = None         # parallel suggestion; cord_type unchanged
    reclassified_reading: Optional[str] = None
    reclassified_reason: Optional[str] = None
    # Pass 3C -------------------------------------------------------------
    constraint_source: Optional[str] = None              # sum-cord that constrained this value
    # Pass 3D (iterative arithmetic repair) --------------------------------
    repaired_turn: Optional[int] = None                  # inferred L-turn count for a damaged L?
    repaired_reading: Optional[str] = None               # syllable derived from repaired_turn
    repaired_via_sum: Optional[str] = None               # sum-cord that enabled the repair
    repaired_iteration: Optional[int] = None             # fixed-point iteration at which it was solved

    def to_dict(self) -> dict:
        """Compact JSON-friendly dict: only emits fields that carry signal."""
        d: dict = {
            "role": self.role,
            "cascade_depth": self.cascade_depth,
            "s_value_total": self.s_value_total,
        }
        if self.expected_locke is not None:
            d["expected_locke"] = self.expected_locke
            d["sum_hand"] = self.sum_hand
            d["summands"] = list(self.summand_pendant_names)
        if self.referenced_by_pendant_names:
            d["referenced_by"] = list(self.referenced_by_pendant_names)
        if self.verified is not None:
            d["verified"] = bool(self.verified)
        if self.decomposition:
            d["decomposition"] = self.decomposition
        if self.reclassified_cord_type:
            d["reclassified_cord_type"] = self.reclassified_cord_type
            d["reclassified_reading"] = self.reclassified_reading
            d["reclassified_reason"] = self.reclassified_reason
        if self.constraint_source:
            d["constraint_source"] = self.constraint_source
        if self.repaired_turn is not None:
            d["repaired_turn"] = self.repaired_turn
            d["repaired_reading"] = self.repaired_reading
            d["repaired_via_sum"] = self.repaired_via_sum
            d["repaired_iteration"] = self.repaired_iteration
        return d


@dataclass
class SumRelation:
    """A single sum relation, resolved against ``KhipuData``.

    ``sum_cord`` and ``summands`` are ``p{N}`` pendant names. If a sum or
    summand could not be resolved in the xlsx, the relation is dropped
    (never included in an AscherGraph).
    """
    hand: str                       # 'right' | 'left'
    sum_cord: str                   # e.g. 'p7'
    sum_value: float                # expected Locke total (from KFG HTML)
    summands: List[str] = field(default_factory=list)
    summand_values: List[float] = field(default_factory=list)


class AscherGraph:
    """Structural summary of a khipu's Ascher pendant-pendant sums.

    The graph is a directed acyclic graph (one node per involved cord,
    one edge per summand → sum-cord relation). Use :meth:`from_kfg` to
    build from KFG data via a :class:`KFGClient`, or the constructor
    directly if you already have ``KhipuData`` + ``List[AscherSum]``
    in hand.
    """

    def __init__(self, kh_id: str, khipu_data: KhipuData,
                 sums: List[AscherSum]):
        if not HAS_NX:
            raise RuntimeError(
                "AscherGraph requires the 'networkx' package. "
                "Install it with: pip install networkx"
            )
        self.kh_id = kh_id
        self.khipu_data = khipu_data

        # Resolved relations (sum-cord name + summand names in p{N} form)
        self.relations: List[SumRelation] = []
        # Indexes
        self._sum_of: Dict[str, SumRelation] = {}     # sum_cord -> relation
        self._referenced_by: Dict[str, List[str]] = defaultdict(list)
        self.graph = nx.DiGraph()

        self._build(sums)

    # -- Construction ----------------------------------------------------
    def _build(self, sums: List[AscherSum]) -> None:
        kd = self.khipu_data
        for s in sums:
            sum_row = resolve_cord(kd, s.sum_group, s.sum_pos)
            if sum_row is None:
                continue
            sum_name = str(sum_row["_cord_name"])
            summand_names: List[str] = []
            summand_values: List[float] = []
            for g, p, v in s.summands:
                r = resolve_cord(kd, g, p)
                if r is None:
                    continue
                summand_names.append(str(r["_cord_name"]))
                summand_values.append(float(v))
            if not summand_names:
                continue
            rel = SumRelation(
                hand=s.hand, sum_cord=sum_name, sum_value=float(s.sum_value),
                summands=summand_names, summand_values=summand_values,
            )
            self.relations.append(rel)
            # A cord may be the sum-cord of multiple sums (rare). We keep
            # only the first here — subsequent ones are still reachable
            # via ``relations`` and the graph edges.
            self._sum_of.setdefault(sum_name, rel)
            for name in summand_names:
                self._referenced_by[name].append(sum_name)
            # Edges: summand -> sum_cord
            self.graph.add_node(sum_name)
            for name in summand_names:
                self.graph.add_edge(name, sum_name)

    @classmethod
    def from_kfg(cls, kh_id: str,
                 client: Optional["KFGClient"] = None) -> "AscherGraph":
        """Download (or load from cache) and build an AscherGraph for ``kh_id``."""
        if client is None:
            client = KFGClient()
        xlsx = client.fetch_xlsx(kh_id)
        html = client.fetch_sums_html(kh_id)
        kd = parse_khipu_xlsx(xlsx, kh_id)
        sums = parse_sums_html(html, kh_id)
        return cls(kh_id, kd, sums)

    # -- Query API -------------------------------------------------------
    def get_role(self, cord_name: str) -> str:
        """Return HEADER / DATA / BOTH / FREE for a pendant name."""
        is_sum = cord_name in self._sum_of
        is_summand = cord_name in self._referenced_by
        if is_sum and is_summand:
            return ROLE_BOTH
        if is_sum:
            return ROLE_HEADER
        if is_summand:
            return ROLE_DATA
        return ROLE_FREE

    def get_summands(self, sum_cord: str) -> List[str]:
        """Return summand pendant names for a given sum cord (empty if none)."""
        rel = self._sum_of.get(sum_cord)
        return list(rel.summands) if rel else []

    def get_sum_value(self, sum_cord: str) -> Optional[float]:
        """Return the expected Locke total for a sum cord, or None."""
        rel = self._sum_of.get(sum_cord)
        return rel.sum_value if rel else None

    def get_sum_hand(self, sum_cord: str) -> Optional[str]:
        rel = self._sum_of.get(sum_cord)
        return rel.hand if rel else None

    def get_sums_referencing(self, cord_name: str) -> List[str]:
        """Return the sum-cord names that reference ``cord_name`` as a summand."""
        return list(self._referenced_by.get(cord_name, ()))

    def get_cascade_depth(self, cord_name: str) -> int:
        """Return the longest upward path length from this cord in the DAG.

        0 = leaf (DATA cord not referenced anywhere above), higher values
        = closer to the root of the Merkle tree. Isolated cords return 0.
        """
        if cord_name not in self.graph:
            return 0
        # Depth from this node to any sink. networkx has no ready-made
        # "descendant longest path length", so we walk iteratively.
        visited: Dict[str, int] = {}

        def depth(n: str) -> int:
            if n in visited:
                return visited[n]
            successors = list(self.graph.successors(n))
            if not successors:
                visited[n] = 0
                return 0
            d = 1 + max(depth(s) for s in successors)
            visited[n] = d
            return d

        return depth(cord_name)

    def max_cascade_depth(self) -> int:
        """Longest path in the DAG (0 for an empty or flat graph)."""
        if len(self.graph) == 0:
            return 0
        if not nx.is_directed_acyclic_graph(self.graph):
            return -1
        return nx.dag_longest_path_length(self.graph)

    # -- Integrity check -------------------------------------------------
    def verify_sum(self, sum_cord: str,
                   cord_values: Dict[str, float]) -> Optional[bool]:
        """Does the Locke total of summand values match the expected sum value?

        ``cord_values`` maps pendant name -> Locke value (int or float).
        Returns True/False, or None if the sum cord is not known or any
        summand has no value.
        """
        rel = self._sum_of.get(sum_cord)
        if rel is None:
            return None
        try:
            actual = sum(cord_values[name] for name in rel.summands)
        except KeyError:
            return None
        return abs(actual - rel.sum_value) < 1e-9

    # -- OKR bridge / cross-source sanity check --------------------------
    def iter_cord_names(self) -> List[str]:
        """Return every pendant name that is part of at least one sum relation."""
        names: "set[str]" = set()
        for rel in self.relations:
            names.add(rel.sum_cord)
            names.update(rel.summands)
        return sorted(names, key=lambda n: int(n[1:]) if n[1:].isdigit() else 1_000_000)

    def validate_mapping(self, okr_cord_values: Dict[str, float],
                         tolerance: float = 0.05) -> Tuple[float, List[str]]:
        """Compare KFG Locke values against an OKR-sourced mapping.

        Parameters
        ----------
        okr_cord_values : dict
            Mapping pendant name (``p1``, ``p2`` …) to the Locke value the
            caller has computed from its own (OKR / SQLite) source.
        tolerance : float
            Maximum acceptable fraction of mismatched cords (default 5%).

        Returns
        -------
        (mismatch_ratio, mismatches) :
            ``mismatch_ratio`` is mismatches / n_checked (0.0 to 1.0).
            ``mismatches`` lists pendant names where the two sources differ.
            An empty ``mismatches`` means every cord present in both
            sources agrees.

        Notes
        -----
        Only cords present in both sources are compared. Cords missing
        from one side are not counted as mismatches (typical cause:
        subsidiary cords not exposed by KFG).
        """
        kfg_values = {}
        if not self.khipu_data.cords.empty and "Value" in self.khipu_data.cords.columns:
            for _, row in self.khipu_data.cords.iterrows():
                name = row.get("Cord_Name")
                val = row.get("Value")
                if isinstance(name, str) and val is not None:
                    try:
                        kfg_values[name] = float(val)
                    except (TypeError, ValueError):
                        pass

        mismatches: List[str] = []
        n_checked = 0
        for name, okr_val in okr_cord_values.items():
            if name not in kfg_values:
                continue
            n_checked += 1
            if abs(kfg_values[name] - float(okr_val)) > 1e-9:
                mismatches.append(name)
        ratio = (len(mismatches) / n_checked) if n_checked else 0.0
        return ratio, mismatches


# ---------------------------------------------------------------------------
# Wrapper — attach Ascher annotations to a TranslationResult
# ---------------------------------------------------------------------------
#
# This wrapper is the "Pass 2" entry point described in the v2 brief. It
# augments an existing TranslationResult with per-cord Ascher metadata
# (role, cascade depth, expected Locke, summand list, ...). It does NOT:
#   - modify any existing CordTranslation field
#   - change cord_type, alba_reading, locke_value, or confidences
#   - touch document-type detection or vocabulary
#
# Passes 3A (reclassification), 3B (verification + decomposition) and
# 3C (constraint propagation) are layered on top in subsequent commits.


def apply_ascher_constraints(
    result,
    graph: "AscherGraph",
    *,
    verify: bool = True,
) -> object:
    """Enrich ``result`` in place with per-cord Ascher annotations.

    Parameters
    ----------
    result : TranslationResult
        A v1 translation produced by :func:`khipu_translator.translator.translate`.
    graph : AscherGraph
        Pre-built Ascher graph for the same khipu.
    verify : bool
        If True (default), also runs Pass 3B's checksum verification:
        sets ``annotation.verified`` for each HEADER / BOTH cord based on
        whether the sum of its summands' reader-computed Locke values
        equals the KFG-declared sum total.

    Returns
    -------
    TranslationResult
        The same object (mutated). Cords that are not part of any sum
        keep ``cord.ascher = None``.

    Notes
    -----
    Only level-1 cords (primary-cord pendants) are eligible for
    annotations — KFG pendant-pendant sums do not address subsidiary cords.
    The mapping uses ``CordTranslation.global_ordinal`` ↔ KFG ``p{N}``.
    """
    # Build a lookup {pendant_name: CordTranslation} for primary cords only.
    by_pendant: Dict[str, object] = {
        f"p{c.global_ordinal}": c for c in result.cords if c.level == 1
    }

    # Reader-computed Locke totals (from knot_sequence strings). These are
    # used instead of ``cord.locke_value`` because the v1 translator leaves
    # locke_value=None on STRING cords.
    reader_values: Dict[str, int] = {
        p: _locke_total_from_seq(c.knot_sequence) for p, c in by_pendant.items()
    }

    for pname, cord in by_pendant.items():
        role = graph.get_role(pname)
        if role == ROLE_FREE:
            continue
        ann = AscherCordAnnotation(
            role=role,
            cascade_depth=graph.get_cascade_depth(pname),
            s_value_total=_s_padding_from_seq(cord.knot_sequence),
        )
        if role in (ROLE_HEADER, ROLE_BOTH):
            ann.expected_locke = graph.get_sum_value(pname)
            ann.sum_hand = graph.get_sum_hand(pname)
            ann.summand_pendant_names = graph.get_summands(pname)
            if verify and ann.expected_locke is not None:
                summand_total = sum(
                    reader_values.get(s, 0) for s in ann.summand_pendant_names
                )
                my_total = reader_values.get(pname, 0)
                ann.verified = (
                    abs(summand_total - ann.expected_locke) < 1e-9
                    and abs(my_total - ann.expected_locke) < 1e-9
                )
            # Pass 3B decomposition surface: when the reader produced a
            # compound (space-separated root + suffix chain), echo it in
            # the Ascher annotation so consumers can tell at a glance
            # whether morphological decomposition was applied.
            compound = getattr(cord, "alba_compound", None)
            if compound:
                ann.decomposition = str(compound)
        if role in (ROLE_DATA, ROLE_BOTH):
            ann.referenced_by_pendant_names = graph.get_sums_referencing(pname)

        cord.ascher = ann
    return result


# HIGH-confidence L-turn counts (per syllabary.py v3): L0, L2..L9.
# L1 is eliminated, L10 is MEDIUM, L11/L12 are LOW.
_HIGH_CONF_TURNS = {0, 2, 3, 4, 5, 6, 7, 8, 9}
# Pass 3A reclassification threshold: ≥ this fraction of co-summands must
# be STRING for an INT candidate to be reclassified. 2/3 is strict.
_RECLASSIFY_STRING_FRAC = 2 / 3


def _single_l_turn_count(seq: Optional[str]) -> Optional[int]:
    """Return the turn count if knot_sequence is exactly one L{N} and nothing
    else; otherwise None. Used by Pass 3A reclassification."""
    toks = [t for t in (seq or "").split() if t.strip()]
    if len(toks) != 1:
        return None
    t = toks[0]
    if not (t.startswith("L") and t[1:].isdigit()):
        return None
    return int(t[1:])


def apply_reclassification(result, graph: "AscherGraph") -> object:
    """Pass 3A — suggest INT→STRING reclassification in STRING context.

    An INT cord that is a summand (DATA or BOTH role) gets a *parallel*
    STRING suggestion when all three guardrails hold:

      1. Its knot_sequence is a single L{N} token (no S/E padding).
      2. ``N`` is in the HIGH-confidence turn set (L0, L2..L9).
      3. At least :data:`_RECLASSIFY_STRING_FRAC` of its co-summands in
         at least one sum it participates in are STRING.

    The suggestion is written to ``cord.ascher.reclassified_cord_type``,
    ``reclassified_reading`` and ``reclassified_reason``. The original
    ``cord_type`` is NOT mutated.

    Must be called AFTER :func:`apply_ascher_constraints` (which sets up
    the annotation scaffold). Idempotent; safe to call repeatedly.
    """
    # Local import to avoid a cycle if translator re-exports ascher types.
    from khipu_translator.syllabary import TURNS_TO_ONSET

    by_pendant: Dict[str, object] = {
        f"p{c.global_ordinal}": c for c in result.cords if c.level == 1
    }

    n_applied = 0
    for pname, cord in by_pendant.items():
        ann = getattr(cord, "ascher", None)
        if ann is None or ann.role not in (ROLE_DATA, ROLE_BOTH):
            continue
        if cord.cord_type != "INT":
            continue
        turns = _single_l_turn_count(cord.knot_sequence)
        if turns is None or turns not in _HIGH_CONF_TURNS:
            continue

        # Guardrail 3: at least one parent sum where siblings are ≥2/3 STRING.
        accepted = False
        strongest_frac = 0.0
        strongest_parent = None
        for sum_cord_name in ann.referenced_by_pendant_names:
            siblings = [
                by_pendant[s] for s in graph.get_summands(sum_cord_name)
                if s != pname and s in by_pendant
            ]
            if not siblings:
                continue
            n_string_sib = sum(1 for s in siblings if s.cord_type == "STRING")
            frac = n_string_sib / len(siblings)
            if frac > strongest_frac:
                strongest_frac = frac
                strongest_parent = sum_cord_name
            if frac >= _RECLASSIFY_STRING_FRAC:
                accepted = True

        if not accepted:
            continue

        # Assign the suggestion (parallel, non-destructive).
        ann.reclassified_cord_type = "STRING"
        ann.reclassified_reading = TURNS_TO_ONSET.get(turns, f"[L{turns}?]")
        ann.reclassified_reason = (
            f"single L-knot in STRING context "
            f"(parent {strongest_parent}: {int(strongest_frac*100)}% STRING siblings)"
        )
        n_applied += 1

    # Attach a summary stat on the result for the report writer.
    try:
        result.stats["ascher_reclassified"] = n_applied
    except Exception:
        pass
    return result


def apply_pure_arithmetic_repair(
    result, graph: "AscherGraph", *,
    max_iter: int = 30,
) -> object:
    """Pass 3D — iterative fixed-point repair of single-L? damaged cords,
    using ONLY Ascher sum constraints (no external cord-total source).

    This is the "self-correcting code" demonstration: we treat each L?
    cord as an unknown, each sum as a linear constraint, and iterate:

        while some cord was repaired in the previous pass:
            for each unrepaired single-L? cord `c`:
                for each sum that references `c`:
                    if all OTHER summands of that sum have known values:
                        derive c's total = sum_value − Σ(other summands)
                        infer L-turn = c.total − c.S_prefix − c.E_count
                                                − c.known_L_contributions
                        if 0 or 2..9: mark c as repaired, record turn + via_sum

    Analogous to belief-propagation decoding of LDPC codes: each repair
    reduces the unknown count in neighbouring sums, which may then become
    solvable. Converges when no further repair is possible — i.e. we hit
    the erasure-correction capacity of the code under current damage.

    Archaeologically honest: uses nothing outside the khipu itself (no KFG
    Value column, no master-record). If a cord is in NO sum, it cannot be
    repaired. If multiple L? cords live in the same sum without independent
    constraints, the system is unsolvable.

    Writes to ``cord.ascher.repaired_turn / _reading / _via_sum /
    _iteration`` on success. Must be called AFTER
    :func:`apply_ascher_constraints`.

    Parameters
    ----------
    max_iter : int
        Safety bound on iterations (default 30). Convergence is usually
        reached within 3–6 iterations.

    Returns
    -------
    TranslationResult
        Same object, mutated. ``result.stats`` gains
        ``ascher_repaired`` (total count), ``ascher_repair_iterations``
        (iterations to convergence), and ``ascher_repair_unresolved``
        (remaining L? cords that cannot be repaired).
    """
    from khipu_translator.syllabary import TURNS_TO_ONSET, TURNS_TO_SYLLABLE

    # HIGH-confidence L-turn set (same as Pass 3A)
    HIGH_CONF = {0, 2, 3, 4, 5, 6, 7, 8, 9}

    # Build the working set: level-1 cords keyed by pendant number.
    by_pendant: Dict[str, object] = {
        f"p{c.global_ordinal}": c for c in result.cords if c.level == 1
    }

    def _parse_seq(seq: Optional[str]) -> Tuple[int, int, int, int, int]:
        """(s_total, l_known_total, l_unknown_count, e_count, known_cord_value_or_None)."""
        s = l_known = l_unk = e = 0
        for tok in (seq or "").split():
            tok = tok.strip()
            if tok == "E":
                e += 1
            elif tok == "L?":
                l_unk += 1
            elif tok.startswith("L") and tok[1:].isdigit():
                l_known += int(tok[1:])
            elif tok.startswith("S") and tok[1:].isdigit():
                s += int(tok[1:])
        return s, l_known, l_unk, e, (s + l_known + e if l_unk == 0 else None)

    # Initial state: known_value[pname] = cord's full Locke value if no L?
    known_value: Dict[str, int] = {}
    unknown_singles: "set[str]" = set()     # 1 L? and no other damage
    for pname, cord in by_pendant.items():
        s, lk, lu, e, full = _parse_seq(cord.knot_sequence)
        if lu == 0:
            known_value[pname] = full
        elif lu == 1:
            unknown_singles.add(pname)
        # Multi-L? cords are not handled by Pass 3D (we'd need N-linear solving).

    # Precompute per-cord (s, lk, lu, e) to avoid repeated parsing inside the loop.
    cord_parts: Dict[str, Tuple[int, int, int, int]] = {
        pname: _parse_seq(cord.knot_sequence)[:4]
        for pname, cord in by_pendant.items()
    }

    iteration = 0
    total_repaired = 0
    while iteration < max_iter:
        iteration += 1
        repaired_this_pass = 0
        # Take a snapshot of the unknowns we'll try this pass.
        for pname in list(unknown_singles):
            cord = by_pendant[pname]
            ann = getattr(cord, "ascher", None)
            if ann is None or ann.role == ROLE_FREE:
                continue
            s, lk, lu, e = cord_parts[pname]

            # Collect candidate constraints (cord.total hints) from every
            # incident sum relation. Two kinds:
            #   (a) HEADER / BOTH: cord is the sum-cord of some relation,
            #       so its declared total = that relation's sum_value.
            #   (b) DATA / BOTH:   cord is a summand, so its total =
            #       parent_sum_value − Σ(other summands' known values).
            # We accept the first constraint that yields a valid HIGH_CONF
            # integer turn; belief-propagation style.
            inferred_turn = None
            via_sum = None

            # (a) HEADER-side constraint
            if ann.role in (ROLE_HEADER, ROLE_BOTH) and ann.expected_locke is not None:
                candidate_total = float(ann.expected_locke)
                candidate_turn = candidate_total - s - lk - e
                if float(candidate_turn).is_integer():
                    candidate_turn = int(candidate_turn)
                    if candidate_turn in HIGH_CONF:
                        inferred_turn = candidate_turn
                        # Record the sum this cord heads as the source.
                        via_sum = pname  # it's its own sum-cord
                        # Also find the relation for provenance display
                        for rel in graph.relations:
                            if rel.sum_cord == pname:
                                via_sum = rel.sum_cord
                                break

            # (b) DATA-side: iterate parent sums
            if inferred_turn is None and ann.role in (ROLE_DATA, ROLE_BOTH):
                for sum_cord_name in ann.referenced_by_pendant_names:
                    rel = graph._sum_of.get(sum_cord_name)
                    if rel is None:
                        continue
                    others_ok = True
                    others_total = 0
                    for sname in rel.summands:
                        if sname == pname:
                            continue
                        if sname not in known_value:
                            others_ok = False
                            break
                        others_total += known_value[sname]
                    if not others_ok:
                        continue
                    candidate_total = rel.sum_value - others_total
                    candidate_turn = candidate_total - s - lk - e
                    if not float(candidate_turn).is_integer():
                        continue
                    candidate_turn = int(candidate_turn)
                    if candidate_turn not in HIGH_CONF:
                        continue
                    inferred_turn = candidate_turn
                    via_sum = sum_cord_name
                    break

            if inferred_turn is None:
                continue

            # Determine syllable. Onset polyphony applies when the L? is
            # the first L-knot in the cord (no known L before it).
            is_onset = (lk == 0)
            table = TURNS_TO_ONSET if is_onset else TURNS_TO_SYLLABLE
            syllable = table.get(inferred_turn, f"L{inferred_turn}?")

            ann.repaired_turn = inferred_turn
            ann.repaired_reading = str(syllable)
            ann.repaired_via_sum = via_sum
            ann.repaired_iteration = iteration

            known_value[pname] = s + lk + e + inferred_turn
            unknown_singles.discard(pname)
            total_repaired += 1
            repaired_this_pass += 1

        if repaired_this_pass == 0:
            break  # fixed point reached

    try:
        result.stats["ascher_repaired"] = total_repaired
        result.stats["ascher_repair_iterations"] = iteration
        result.stats["ascher_repair_unresolved"] = len(unknown_singles)
    except Exception:
        pass
    return result


def apply_constraint_propagation(result, graph: "AscherGraph") -> object:
    """Pass 3C — propagate sum constraints to exactly-one-unknown summands.

    For each sum relation, let the *unknown* summands be those whose
    reader-computed Locke value disagrees with the KFG-declared summand
    value (or whose Locke total we simply couldn't compute, e.g. due to
    an unreadable L? token). If exactly one summand is unknown, we can
    derive its expected value arithmetically:

        expected_unknown = expected_sum − sum(known summand values)

    When the derived value matches the KFG-declared summand value for
    that slot, we annotate the unknown cord with ``constraint_source``
    (the sum cord that uniquely determined it). We do NOT mutate the
    reading or value — this is an attention marker for the researcher.

    Must be called AFTER :func:`apply_ascher_constraints`. Idempotent.
    """
    by_pendant: Dict[str, object] = {
        f"p{c.global_ordinal}": c for c in result.cords if c.level == 1
    }
    reader_values: Dict[str, int] = {
        p: _locke_total_from_seq(c.knot_sequence) for p, c in by_pendant.items()
    }

    n_applied = 0
    for rel in graph.relations:
        # Determine unknowns vs knowns using the KFG-declared summand values
        # as ground truth. An "unknown" here is a summand whose reader-computed
        # value disagrees with KFG — i.e., we couldn't parse it cleanly.
        unknowns: List[Tuple[str, float]] = []   # (name, declared_value)
        known_total = 0.0
        resolved_ok = True
        for name, declared in zip(rel.summands, rel.summand_values):
            if name not in by_pendant:
                resolved_ok = False
                break
            got = reader_values.get(name, 0)
            if abs(got - declared) < 1e-9:
                known_total += declared
            else:
                unknowns.append((name, declared))
        if not resolved_ok or len(unknowns) != 1:
            continue

        unk_name, declared = unknowns[0]
        derived = rel.sum_value - known_total
        if abs(derived - declared) > 1e-9:
            # The sum doesn't close even with KFG's own values — skip.
            continue

        cord = by_pendant[unk_name]
        ann = getattr(cord, "ascher", None)
        if ann is None:
            # Can happen if the cord got a FREE role earlier; create the
            # annotation on the fly so we don't silently drop the signal.
            ann = AscherCordAnnotation(
                role=ROLE_DATA,
                cascade_depth=graph.get_cascade_depth(unk_name),
                s_value_total=_s_padding_from_seq(cord.knot_sequence),
                referenced_by_pendant_names=graph.get_sums_referencing(unk_name),
            )
            cord.ascher = ann
        if ann.constraint_source is None:
            ann.constraint_source = rel.sum_cord
            n_applied += 1

    try:
        result.stats["ascher_propagated"] = n_applied
    except Exception:
        pass
    return result


def _locke_total_from_seq(seq: Optional[str]) -> int:
    """Reader-side Locke total from a knot_sequence string.

    The reader emits tokens ``S{value}`` (simple-knot cluster),
    ``L{turns}`` (long knot) and ``E`` (figure-eight) separated by spaces.
    Unreadable turn counts appear as ``L?`` and are skipped.
    """
    total = 0
    for tok in (seq or "").split():
        tok = tok.strip()
        if not tok:
            continue
        if tok == "E":
            total += 1
        elif tok.startswith("L") and tok[1:].isdigit():
            total += int(tok[1:])
        elif tok.startswith("S") and tok[1:].isdigit():
            total += int(tok[1:])
    return total


def _s_padding_from_seq(seq: Optional[str]) -> int:
    """Sum of S-knot cluster values (scribe padding total)."""
    total = 0
    for tok in (seq or "").split():
        tok = tok.strip()
        if tok.startswith("S") and tok[1:].isdigit():
            total += int(tok[1:])
    return total


